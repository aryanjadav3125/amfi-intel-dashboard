import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelGenerator:
    """
    Generates structured spreadsheet reports for Mutual Funds.
    """
    def __init__(self, output_dir: str = "reports_output"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def generate_fund_report(self, fund_data: dict, filename: str) -> str:
        filepath = os.path.join(self.output_dir, filename)
        wb = Workbook()
        
        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        
        ws1['A1'] = "Fund Overview"
        ws1['A1'].font = Font(bold=True, size=14)
        
        ws1['A3'] = "Fund Name"
        ws1['B3'] = fund_data.get('name', 'N/A')
        ws1['A4'] = "Category"
        ws1['B4'] = fund_data.get('category', 'N/A')
        ws1['A5'] = "AMC"
        ws1['B5'] = fund_data.get('amc', 'N/A')
        ws1['A6'] = "Latest NAV"
        ws1['B6'] = fund_data.get('nav', 'N/A')
        
        ws1['D3'] = "Performance"
        ws1['D3'].font = Font(bold=True, size=12)
        ws1['D4'] = "1Y CAGR"
        ws1['E4'] = f"{fund_data.get('cagr_1y', 'N/A')}%"
        ws1['D5'] = "3Y CAGR"
        ws1['E5'] = f"{fund_data.get('cagr_3y', 'N/A')}%"
        ws1['D6'] = "5Y CAGR"
        ws1['E6'] = f"{fund_data.get('cagr_5y', 'N/A')}%"
        
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 20
        
        # Sheet 2: Holdings
        ws2 = wb.create_sheet(title="Holdings")
        headers = ["Company", "Sector", "Allocation (%)"]
        ws2.append(headers)
        
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        holdings = fund_data.get('holdings', [])
        for h in holdings:
            ws2.append([h['company'], h['sector'], h['allocation']])
            
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 15
        
        wb.save(filepath)
        return filepath
